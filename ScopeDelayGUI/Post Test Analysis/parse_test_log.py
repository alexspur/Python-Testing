"""
Enhanced Marx Generator Test Log Parser
- Extracts all test parameters and results
- Tracks trigger attempts, errors, timestamps
- Generates charts and statistical analysis
- Works with reorganized folder structure
"""

import re
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from collections import defaultdict


def auto_fit_columns(ws, min_width=8, max_width=80, padding=2):
    """Automatically adjust column widths to fit content"""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Apply width with padding, respecting min/max limits
        adjusted_width = max(min_width, min(max_length + padding, max_width))
        ws.column_dimensions[column_letter].width = adjusted_width

def parse_timestamp_from_filename(filename):
    """Extract datetime from log filename like experiment_log_20251211_094947.csv"""
    match = re.search(r'(\d{8})_(\d{6})', filename)
    if match:
        date_str = match.group(1)
        time_str = match.group(2)
        return datetime.strptime(f"{date_str}_{time_str}", "%Y%m%d_%H%M%S")
    return None

def parse_test_log(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Split by test sections
    test_pattern = r'(Test\s*\d+[:\s].*?)(?=Test\s*\d+[:\s]|$)'
    test_blocks = re.findall(test_pattern, content, re.DOTALL | re.IGNORECASE)
    
    results = []
    
    for block in test_blocks:
        test_data = parse_single_test(block)
        if test_data:
            results.append(test_data)
    
    return results

def parse_single_test(block):
    """Parse a single test block and extract all relevant data"""
    
    # Extract test number
    test_num_match = re.search(r'Test\s*(\d+)', block, re.IGNORECASE)
    if not test_num_match:
        return None
    test_num = int(test_num_match.group(1))
    
    # Extract PSI
    psi_match = re.search(r'(\d+)\s*psi', block, re.IGNORECASE)
    psi = int(psi_match.group(1)) if psi_match else None
    
    # Extract all voltage settings from WJ commands
    voltage_matches = re.findall(r'\[WJ[12]\]\s*Set\s*→\s*(\d+\.?\d*)\s*kV', block)
    voltages = [float(v) for v in voltage_matches]
    voltage_steps = sorted(set(voltages)) if voltages else []
    charging_voltage = max(voltages) if voltages else None
    
    # Extract experiment log filename
    exp_log_match = re.search(r'experiment_log_(\d{8}_\d{6})\.csv', block)
    exp_log = f"experiment_log_{exp_log_match.group(1)}.csv" if exp_log_match else None
    start_time = parse_timestamp_from_filename(exp_log) if exp_log else None
    
    # Extract Rigol CSV filenames
    rigol_files = re.findall(r'(rigol\d_\d{8}_\d{6}\.csv)', block)
    
    # Get end time from last Rigol file
    end_time = None
    if rigol_files:
        last_rigol = rigol_files[-1]
        end_time = parse_timestamp_from_filename(last_rigol)
    
    # Calculate duration
    duration_seconds = None
    if start_time and end_time:
        duration_seconds = (end_time - start_time).total_seconds()
    
    # Count BNC trigger attempts
    bnc_fires = len(re.findall(r'\[BNC575\] Internal pulse fired', block))
    
    # Extract errors and warnings
    errors = []
    error_patterns = [
        (r'\[WJ\d ERROR\]\s*(.+)', 'WJ Error'),
        (r'\[WJ\d\].*?\'type\':\s*\'E\'.*?\'message\':\s*\'([^\']+)\'', 'WJ Command Error'),
        (r'NOT CONNECTED:\s*(.+)', 'Connection Error'),
    ]
    for pattern, error_type in error_patterns:
        matches = re.findall(pattern, block)
        for match in matches:
            errors.append(f"{error_type}: {match[:50]}")
    
    # Count connection issues
    connection_retries = len(re.findall(r'\[WJ\d\] Connecting on COM', block))
    disconnects = len(re.findall(r'\[WJ\d\] Disconnected', block))
    timeouts = len(re.findall(r'Write timeout', block))
    
    # Determine if triggered or self-broke
    lines = block.split('\n')
    events = []
    for line in lines:
        if '[BNC575] Internal pulse fired' in line:
            events.append('bnc_fire')
        if '[Rigol #' in line and 'capture complete' in line:
            events.append('rigol_complete')
        if '[WJ' in line and 'Set →' in line:
            events.append('voltage_set')
    
    # Triggered: BNC fires and Rigol completes shortly after without voltage changes
    triggered = False
    for i, event in enumerate(events):
        if event == 'bnc_fire':
            remaining = events[i+1:i+5]
            if 'rigol_complete' in remaining:
                voltage_sets_before_rigol = 0
                for e in remaining:
                    if e == 'voltage_set':
                        voltage_sets_before_rigol += 1
                    if e == 'rigol_complete':
                        break
                if voltage_sets_before_rigol == 0:
                    triggered = True
                    break
    
    # Override with explicit mentions in the test description
    header_lines = block.split('===')[0] if '===' in block else block[:500]
    if 'self broke' in header_lines.lower() or 'self-break' in header_lines.lower():
        if 'no fire' in header_lines.lower() or 'attempted trigger' in header_lines.lower():
            triggered = False
        elif 'trigger' not in header_lines.lower():
            triggered = False
    elif 'trigger' in header_lines.lower() and 'no fire' not in header_lines.lower():
        if 'attempted trigger' not in header_lines.lower():
            triggered = True
    
    # Extract negative and positive power supply readings
    neg_ps_match = re.search(r'[Nn]eg(?:ative)?[:\s]*(\d+\.?\d*)\s*k?V', block)
    pos_ps_match = re.search(r'[Pp]os(?:itive)?[:\s]*(\d+\.?\d*)\s*k?V', block)
    neg_ps = float(neg_ps_match.group(1)) if neg_ps_match else None
    pos_ps = float(pos_ps_match.group(1)) if pos_ps_match else None
    
    # Get description/notes from the test header
    description = ""
    if '===' in block:
        header = block.split('===')[0].strip()
        desc_match = re.search(r'Test\s*\d+[:\s]*(.*)', header, re.DOTALL)
        if desc_match:
            description = desc_match.group(1).strip()
            description = ' '.join(description.split())
    
    # Generate folder name (matching your structure)
    safe_desc = re.sub(r'[<>:"/\\|?*]', '', description)[:60] if description else ""
    safe_desc = safe_desc.replace(' ', '_')
    folder_name = f"Test_{test_num:02d}_{safe_desc}" if safe_desc else f"Test_{test_num:02d}"
    
    return {
        'test_num': test_num,
        'psi': psi,
        'charging_voltage_kV': charging_voltage,
        'voltage_steps': voltage_steps,
        'triggered': triggered,
        'status': 'Triggered' if triggered else 'Self-Broke',
        'trigger_attempts': bnc_fires,
        'neg_ps_kV': neg_ps,
        'pos_ps_kV': pos_ps,
        'description': description[:100] if description else "",
        'folder_name': folder_name,
        'experiment_log': exp_log,
        'rigol_files': rigol_files,
        'start_time': start_time,
        'end_time': end_time,
        'duration_seconds': duration_seconds,
        'errors': errors,
        'timeouts': timeouts,
        'disconnects': disconnects,
        'connection_retries': connection_retries,
    }

def calculate_statistics(results):
    """Calculate statistics grouped by PSI"""
    stats_by_psi = defaultdict(lambda: {
        'voltages': [],
        'triggered_count': 0,
        'self_broke_count': 0,
        'total': 0
    })
    
    for r in results:
        if r['psi'] is not None and r['charging_voltage_kV'] is not None:
            psi = r['psi']
            stats_by_psi[psi]['voltages'].append(r['charging_voltage_kV'])
            stats_by_psi[psi]['total'] += 1
            if r['triggered']:
                stats_by_psi[psi]['triggered_count'] += 1
            else:
                stats_by_psi[psi]['self_broke_count'] += 1
    
    stats_summary = []
    for psi in sorted(stats_by_psi.keys()):
        data = stats_by_psi[psi]
        voltages = data['voltages']
        avg_voltage = sum(voltages) / len(voltages) if voltages else 0
        min_voltage = min(voltages) if voltages else 0
        max_voltage = max(voltages) if voltages else 0
        std_dev = (sum((v - avg_voltage)**2 for v in voltages) / len(voltages))**0.5 if len(voltages) > 1 else 0
        trigger_rate = (data['triggered_count'] / data['total'] * 100) if data['total'] > 0 else 0
        
        stats_summary.append({
            'psi': psi,
            'num_tests': data['total'],
            'avg_voltage': avg_voltage,
            'min_voltage': min_voltage,
            'max_voltage': max_voltage,
            'std_dev': std_dev,
            'triggered': data['triggered_count'],
            'self_broke': data['self_broke_count'],
            'trigger_success_rate': trigger_rate
        })
    
    return stats_summary

def create_excel(results, stats, output_path):
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    triggered_fill = PatternFill('solid', fgColor='C6EFCE')
    self_broke_fill = PatternFill('solid', fgColor='FFC7CE')
    warning_fill = PatternFill('solid', fgColor='FFEB9C')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ========== SHEET 1: Main Results ==========
    ws1 = wb.active
    ws1.title = "Test Results"
    
    headers1 = ['Test #', 'PSI', 'Voltage (kV)', 'Status', 'Trigger Attempts', 
                'Duration (s)', 'Neg PS (kV)', 'Pos PS (kV)', 'Errors', 'Folder']
    
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, r in enumerate(results, 2):
        ws1.cell(row=row_idx, column=1, value=r['test_num']).border = thin_border
        ws1.cell(row=row_idx, column=2, value=r['psi']).border = thin_border
        ws1.cell(row=row_idx, column=3, value=r['charging_voltage_kV']).border = thin_border
        
        status_cell = ws1.cell(row=row_idx, column=4, value=r['status'])
        status_cell.border = thin_border
        status_cell.fill = triggered_fill if r['triggered'] else self_broke_fill
        
        ws1.cell(row=row_idx, column=5, value=r['trigger_attempts']).border = thin_border
        ws1.cell(row=row_idx, column=6, value=r['duration_seconds']).border = thin_border
        ws1.cell(row=row_idx, column=7, value=r['neg_ps_kV']).border = thin_border
        ws1.cell(row=row_idx, column=8, value=r['pos_ps_kV']).border = thin_border
        
        error_count = len(r['errors']) + r['timeouts']
        error_cell = ws1.cell(row=row_idx, column=9, value=error_count)
        error_cell.border = thin_border
        if error_count > 0:
            error_cell.fill = warning_fill
        
        ws1.cell(row=row_idx, column=10, value=r['folder_name']).border = thin_border
    
    # Auto-fit column widths
    auto_fit_columns(ws1)
    
    # ========== SHEET 2: Statistics by PSI ==========
    ws2 = wb.create_sheet("Statistics by PSI")
    
    headers2 = ['PSI', '# Tests', 'Avg Voltage (kV)', 'Min (kV)', 'Max (kV)', 
                'Std Dev', 'Triggered', 'Self-Broke', 'Trigger Rate (%)']
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, s in enumerate(stats, 2):
        ws2.cell(row=row_idx, column=1, value=s['psi']).border = thin_border
        ws2.cell(row=row_idx, column=2, value=s['num_tests']).border = thin_border
        ws2.cell(row=row_idx, column=3, value=round(s['avg_voltage'], 1)).border = thin_border
        ws2.cell(row=row_idx, column=4, value=s['min_voltage']).border = thin_border
        ws2.cell(row=row_idx, column=5, value=s['max_voltage']).border = thin_border
        ws2.cell(row=row_idx, column=6, value=round(s['std_dev'], 2)).border = thin_border
        ws2.cell(row=row_idx, column=7, value=s['triggered']).border = thin_border
        ws2.cell(row=row_idx, column=8, value=s['self_broke']).border = thin_border
        ws2.cell(row=row_idx, column=9, value=round(s['trigger_success_rate'], 1)).border = thin_border
    
    # Auto-fit column widths
    auto_fit_columns(ws2)
    
    # ========== SHEET 3: Voltage vs PSI Chart Data ==========
    ws3 = wb.create_sheet("Chart Data")
    
    # Get unique PSI values and assign colors
    psi_values = sorted(set(r['psi'] for r in results if r['psi'] is not None))
    
    # Color palette for different PSI values (hex colors)
    psi_colors = [
        'FF4136',  # Red
        'FF851B',  # Orange  
        'FFDC00',  # Yellow
        '2ECC40',  # Green
        '0074D9',  # Blue
        'B10DC9',  # Purple
        'F012BE',  # Magenta
        '39CCCC',  # Teal
        '01FF70',  # Lime
        '85144b',  # Maroon
    ]
    
    psi_color_map = {psi: psi_colors[i % len(psi_colors)] for i, psi in enumerate(psi_values)}
    
    # Headers
    ws3['A1'] = 'PSI'
    ws3['B1'] = 'Breakdown Voltage (kV)'
    ws3['C1'] = 'Status'
    ws3['D1'] = 'Test #'
    for col in 'ABCD':
        ws3[f'{col}1'].fill = header_fill
        ws3[f'{col}1'].font = header_font
        ws3[f'{col}1'].border = thin_border
    
    # Group data by PSI for the chart - each PSI gets its own columns
    # Starting at column F for chart data (E is spacer)
    col_offset = 6  # Column F
    psi_col_map = {}  # Maps PSI to (x_col, y_col)
    
    # Create headers for each PSI
    for i, psi in enumerate(psi_values):
        x_col = col_offset + (i * 2)
        y_col = x_col + 1
        psi_col_map[psi] = (x_col, y_col)
        
        ws3.cell(row=1, column=x_col, value=f'{psi} PSI (X)')
        ws3.cell(row=1, column=y_col, value=f'{psi} PSI (kV)')
        ws3.cell(row=1, column=x_col).fill = header_fill
        ws3.cell(row=1, column=y_col).fill = header_fill
        ws3.cell(row=1, column=x_col).font = header_font
        ws3.cell(row=1, column=y_col).font = header_font
    
    # Fill in main data table and grouped data for chart
    row = 2
    psi_row_counters = {psi: 2 for psi in psi_values}  # Track row for each PSI column
    
    for r in results:
        if r['psi'] is not None and r['charging_voltage_kV'] is not None:
            # Main data table (columns A-D)
            psi = r['psi']
            color = psi_color_map[psi]
            
            cell_a = ws3.cell(row=row, column=1, value=psi)
            cell_a.border = thin_border
            cell_a.fill = PatternFill('solid', fgColor=color)
            
            cell_b = ws3.cell(row=row, column=2, value=r['charging_voltage_kV'])
            cell_b.border = thin_border
            cell_b.fill = PatternFill('solid', fgColor=color)
            
            ws3.cell(row=row, column=3, value=r['status']).border = thin_border
            ws3.cell(row=row, column=4, value=r['test_num']).border = thin_border
            
            # Grouped data for chart (each PSI in its own column pair)
            x_col, y_col = psi_col_map[psi]
            psi_row = psi_row_counters[psi]
            ws3.cell(row=psi_row, column=x_col, value=psi)
            ws3.cell(row=psi_row, column=y_col, value=r['charging_voltage_kV'])
            psi_row_counters[psi] += 1
            
            row += 1
    
    # Create scatter chart with separate series for each PSI
    chart = ScatterChart()
    chart.title = "Breakdown Voltage vs Pressure"
    chart.x_axis.title = "Pressure (PSI)"
    chart.y_axis.title = "Breakdown Voltage (kV)"
    chart.style = 10
    chart.height = 15
    chart.width = 20
    chart.legend.position = 'r'  # Legend on right
    
    # Add a series for each PSI value
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    
    for psi in psi_values:
        x_col, y_col = psi_col_map[psi]
        max_row = psi_row_counters[psi] - 1
        
        if max_row >= 2:  # Only add if there's data
            x_values = Reference(ws3, min_col=x_col, min_row=2, max_row=max_row)
            y_values = Reference(ws3, min_col=y_col, min_row=2, max_row=max_row)
            
            series = Series(y_values, x_values, title=f"{psi} PSI")
            series.marker = Marker(symbol='circle', size=12)
            series.marker.graphicalProperties.solidFill = psi_color_map[psi]
            series.marker.graphicalProperties.line.solidFill = psi_color_map[psi]
            series.graphicalProperties.line.noFill = True
            chart.series.append(series)
    
    ws3.add_chart(chart, "A" + str(row + 3))
    
    # Add color legend below the data table
    legend_row = row + 2
    ws3.cell(row=legend_row, column=1, value="Color Legend:").font = Font(bold=True)
    for i, psi in enumerate(psi_values):
        cell = ws3.cell(row=legend_row, column=2+i, value=f"{psi} PSI")
        cell.fill = PatternFill('solid', fgColor=psi_color_map[psi])
        cell.font = Font(bold=True)
    
    # Auto-fit column widths
    auto_fit_columns(ws3)
    
    # ========== SHEET 4: Detailed Info ==========
    ws4 = wb.create_sheet("Detailed Info")
    
    headers4 = ['Test #', 'Description', 'Start Time', 'End Time', 'Duration',
                'Voltage Steps', 'Experiment Log', 'Rigol Files', 'Errors/Warnings']
    
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, r in enumerate(results, 2):
        ws4.cell(row=row_idx, column=1, value=r['test_num']).border = thin_border
        ws4.cell(row=row_idx, column=2, value=r['description'][:80]).border = thin_border
        
        start_str = r['start_time'].strftime('%H:%M:%S') if r['start_time'] else ''
        end_str = r['end_time'].strftime('%H:%M:%S') if r['end_time'] else ''
        duration_str = f"{int(r['duration_seconds'])}s" if r['duration_seconds'] else ''
        
        ws4.cell(row=row_idx, column=3, value=start_str).border = thin_border
        ws4.cell(row=row_idx, column=4, value=end_str).border = thin_border
        ws4.cell(row=row_idx, column=5, value=duration_str).border = thin_border
        
        voltage_steps_str = ' → '.join([f"{v}kV" for v in r['voltage_steps']])
        ws4.cell(row=row_idx, column=6, value=voltage_steps_str).border = thin_border
        
        ws4.cell(row=row_idx, column=7, value=r['experiment_log'] or '').border = thin_border
        
        rigol_str = ', '.join(r['rigol_files'][:3])
        if len(r['rigol_files']) > 3:
            rigol_str += f" (+{len(r['rigol_files'])-3} more)"
        ws4.cell(row=row_idx, column=8, value=rigol_str).border = thin_border
        
        errors_str = '; '.join(r['errors'][:2]) if r['errors'] else ''
        if r['timeouts'] > 0:
            errors_str += f" [{r['timeouts']} timeouts]"
        error_cell = ws4.cell(row=row_idx, column=9, value=errors_str)
        error_cell.border = thin_border
        if errors_str:
            error_cell.fill = warning_fill
    
    # Auto-fit column widths
    auto_fit_columns(ws4)
    
    # ========== SHEET 5: Error Log ==========
    ws5 = wb.create_sheet("Error Log")
    
    headers5 = ['Test #', 'Error Type', 'Details']
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    error_row = 2
    for r in results:
        test_num = r['test_num']
        for error in r['errors']:
            ws5.cell(row=error_row, column=1, value=test_num).border = thin_border
            ws5.cell(row=error_row, column=2, value='Error').border = thin_border
            ws5.cell(row=error_row, column=3, value=error).border = thin_border
            error_row += 1
        if r['timeouts'] > 0:
            ws5.cell(row=error_row, column=1, value=test_num).border = thin_border
            ws5.cell(row=error_row, column=2, value='Timeout').border = thin_border
            ws5.cell(row=error_row, column=3, value=f"{r['timeouts']} write timeout(s)").border = thin_border
            error_row += 1
        if r['disconnects'] > 0:
            ws5.cell(row=error_row, column=1, value=test_num).border = thin_border
            ws5.cell(row=error_row, column=2, value='Disconnect').border = thin_border
            ws5.cell(row=error_row, column=3, value=f"{r['disconnects']} disconnect(s)").border = thin_border
            error_row += 1
    
    # Auto-fit column widths
    auto_fit_columns(ws5)
    
    # ========== SHEET 6: Summary ==========
    ws6 = wb.create_sheet("Summary")
    
    ws6['A1'] = "TEST SESSION SUMMARY"
    ws6['A1'].font = Font(bold=True, size=14)
    
    total_tests = len(results)
    triggered_tests = sum(1 for r in results if r['triggered'])
    self_broke_tests = total_tests - triggered_tests
    total_errors = sum(len(r['errors']) + r['timeouts'] for r in results)
    
    summary_data = [
        ('Total Tests', total_tests),
        ('Triggered', triggered_tests),
        ('Self-Broke', self_broke_tests),
        ('Trigger Success Rate', f"{triggered_tests/total_tests*100:.1f}%" if total_tests > 0 else "N/A"),
        ('', ''),
        ('PSI Range', f"{min(r['psi'] for r in results if r['psi'])} - {max(r['psi'] for r in results if r['psi'])} psi"),
        ('Voltage Range', f"{min(r['charging_voltage_kV'] for r in results if r['charging_voltage_kV']):.0f} - {max(r['charging_voltage_kV'] for r in results if r['charging_voltage_kV']):.0f} kV"),
        ('', ''),
        ('Total Errors/Warnings', total_errors),
        ('Tests with Errors', sum(1 for r in results if r['errors'] or r['timeouts'] > 0)),
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 3):
        ws6.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws6.cell(row=row_idx, column=2, value=value)
    
    # Auto-fit column widths
    auto_fit_columns(ws6)
    
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")

def print_summary(results, stats):
    """Print summary to console"""
    print("\n" + "="*60)
    print("  TEST SESSION SUMMARY")
    print("="*60)
    
    total = len(results)
    triggered = sum(1 for r in results if r['triggered'])
    
    print(f"\nTotal Tests: {total}")
    print(f"  Triggered:  {triggered} ({triggered/total*100:.1f}%)")
    print(f"  Self-Broke: {total-triggered} ({(total-triggered)/total*100:.1f}%)")
    
    print("\n" + "-"*60)
    print("  STATISTICS BY PSI")
    print("-"*60)
    print(f"{'PSI':>6} {'Tests':>6} {'Avg kV':>10} {'Min':>8} {'Max':>8} {'Trig%':>8}")
    print("-"*60)
    
    for s in stats:
        print(f"{s['psi']:>6} {s['num_tests']:>6} {s['avg_voltage']:>10.1f} {s['min_voltage']:>8.0f} {s['max_voltage']:>8.0f} {s['trigger_success_rate']:>7.0f}%")
    
    print("\n" + "-"*60)
    print("  INDIVIDUAL TESTS")
    print("-"*60)
    
    for r in results:
        status_icon = "✓" if r['triggered'] else "✗"
        errors = f" [{len(r['errors'])} err]" if r['errors'] else ""
        print(f"  Test {r['test_num']:2d}: {r['psi']:2d} psi, {r['charging_voltage_kV']:5.0f} kV, {r['status']:10s} {status_icon} (attempts: {r['trigger_attempts']}){errors}")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
    else:
        # Default paths - update these for your system
        input_file = r"C:\Users\ESpurbeck\Desktop\LANL Project\Marx Testing\Testing\12.11.25\Testing 12_11_25.txt"
        output_file = r"C:\Users\ESpurbeck\Desktop\LANL Project\Marx Testing\Testing\12.11.25\Test_Results_Enhanced.xlsx"
    
    print(f"Reading: {input_file}")
    results = parse_test_log(input_file)
    stats = calculate_statistics(results)
    
    print_summary(results, stats)
    
    create_excel(results, stats, output_file)
    print(f"\nExcel file saved: {output_file}")
